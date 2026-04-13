from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from .models import AttendanceSession, AttendanceRecord
from .serializers import AttendanceSessionSerializer, AttendanceRecordSerializer, StudentAttendanceSummarySerializer
from academics.models import Subject, Section
from users.models import CustomUser
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

# ─── FACULTY VIEWS ────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_attendance_session(request):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can create attendance sessions'}, status=status.HTTP_403_FORBIDDEN)

    subject_id = request.data.get('subject')
    section_id = request.data.get('section')
    date = request.data.get('date')
    start_time = request.data.get('start_time')
    topic_covered = request.data.get('topic_covered', '')
    attendance_data = request.data.get('attendance', [])

    if not all([subject_id, section_id, date, start_time]):
        return Response({'error': 'subject, section, date and start_time are required'}, status=status.HTTP_400_BAD_REQUEST)

    subject = get_object_or_404(Subject, id=subject_id)
    section = get_object_or_404(Section, id=section_id)

    # Check if session already exists
    if AttendanceSession.objects.filter(subject=subject, section=section, date=date, start_time=start_time).exists():
        return Response({'error': 'Attendance session already exists for this slot'}, status=status.HTTP_400_BAD_REQUEST)

    # Create session
    session = AttendanceSession.objects.create(
        faculty=request.user,
        subject=subject,
        section=section,
        date=date,
        start_time=start_time,
        topic_covered=topic_covered
    )

    # Create attendance records for each student
    for item in attendance_data:
        student = get_object_or_404(CustomUser, id=item.get('student_id'))
        AttendanceRecord.objects.create(
            session=session,
            student=student,
            status=item.get('status', 'absent')
        )

    serializer = AttendanceSessionSerializer(session)
    return Response({
        'message': 'Attendance marked successfully',
        'session': serializer.data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_section_students(request):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can access this'}, status=status.HTTP_403_FORBIDDEN)

    section_id = request.query_params.get('section_id')
    if not section_id:
        return Response({'error': 'section_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    section = get_object_or_404(Section, id=section_id)
    students = section.students.filter(role='student')

    data = []
    for student in students:
        try:
            enrollment = student.student_profile.enrollment_number
        except:
            enrollment = 'N/A'
        data.append({
            'id': student.id,
            'name': student.get_full_name(),
            'enrollment_number': enrollment,
            'username': student.username
        })

    return Response(data)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def manage_attendance_session(request, session_id):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can access this'}, status=status.HTTP_403_FORBIDDEN)

    session = get_object_or_404(AttendanceSession, id=session_id, faculty=request.user)

    if request.method == 'GET':
        serializer = AttendanceSessionSerializer(session)
        return Response(serializer.data)

    elif request.method == 'PUT':
        attendance_data = request.data.get('attendance', [])
        for item in attendance_data:
            record = get_object_or_404(AttendanceRecord, session=session, student_id=item.get('student_id'))
            record.status = item.get('status', record.status)
            record.save()

        serializer = AttendanceSessionSerializer(session)
        return Response({
            'message': 'Attendance updated successfully',
            'session': serializer.data
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_faculty_sessions(request):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can access this'}, status=status.HTTP_403_FORBIDDEN)

    subject_id = request.query_params.get('subject_id')
    section_id = request.query_params.get('section_id')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    sessions = AttendanceSession.objects.filter(faculty=request.user)

    if subject_id:
        sessions = sessions.filter(subject_id=subject_id)
    if section_id:
        sessions = sessions.filter(section_id=section_id)
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)

    sessions = sessions.order_by('-date')
    serializer = AttendanceSessionSerializer(sessions, many=True)
    return Response(serializer.data)


# ─── STUDENT VIEWS ────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_attendance_summary(request):
    if request.user.role != 'student':
        return Response({'error': 'Only students can access this'}, status=status.HTTP_403_FORBIDDEN)

    try:
        profile = request.user.student_profile
        subjects = Subject.objects.filter(
            semester=profile.semester,
            department__name=profile.branch
        )
    except:
        return Response({'error': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)

    summary = []
    for subject in subjects:
        total_classes = AttendanceSession.objects.filter(subject=subject).count()
        total_present = AttendanceRecord.objects.filter(
            student=request.user,
            session__subject=subject,
            status='present'
        ).count()
        total_absent = total_classes - total_present
        percentage = round((total_present / total_classes * 100), 2) if total_classes > 0 else 0

        summary.append({
            'subject_id': subject.id,
            'subject_name': subject.name,
            'subject_code': subject.code,
            'total_classes': total_classes,
            'total_present': total_present,
            'total_absent': total_absent,
            'attendance_percentage': percentage
        })

    return Response(summary)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_attendance_detail(request):
    if request.user.role != 'student':
        return Response({'error': 'Only students can access this'}, status=status.HTTP_403_FORBIDDEN)

    subject_id = request.query_params.get('subject_id')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    records = AttendanceRecord.objects.filter(student=request.user)

    if subject_id:
        records = records.filter(session__subject_id=subject_id)
    if date_from:
        records = records.filter(session__date__gte=date_from)
    if date_to:
        records = records.filter(session__date__lte=date_to)

    records = records.order_by('-session__date')

    data = []
    for record in records:
        data.append({
            'date': record.session.date,
            'subject': record.session.subject.name,
            'subject_code': record.session.subject.code,
            'status': record.status,
            'topic_covered': record.session.topic_covered,
            'start_time': record.session.start_time,
        })

    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_attendance_excel(request):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can download attendance reports'}, status=status.HTTP_403_FORBIDDEN)

    # Filters
    subject_id = request.query_params.get('subject_id')
    section_id = request.query_params.get('section_id')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    # Get sessions
    sessions = AttendanceSession.objects.filter(faculty=request.user)
    if subject_id:
        sessions = sessions.filter(subject_id=subject_id)
    if section_id:
        sessions = sessions.filter(section_id=section_id)
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)
    sessions = sessions.order_by('date', 'start_time')

    if not sessions.exists():
        return Response({'error': 'No attendance data found for selected filters'}, status=status.HTTP_404_NOT_FOUND)

    # Get all students from these sessions
    student_ids = AttendanceRecord.objects.filter(
        session__in=sessions
    ).values_list('student_id', flat=True).distinct()

    from users.models import CustomUser
    students = CustomUser.objects.filter(id__in=student_ids).order_by('student_profile__enrollment_number')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # ── Styles ──────────────────────────────────────────────
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    subheader_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    present_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
    absent_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    summary_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    title_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFF", end_color="F8FAFF", fill_type="solid")

    white_bold = Font(color="FFFFFF", bold=True, size=11)
    white_normal = Font(color="FFFFFF", bold=False, size=10)
    dark_bold = Font(color="1E3A5F", bold=True, size=10)
    present_font = Font(color="1B4332", bold=True, size=10)
    absent_font = Font(color="7F1D1D", bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ── Title Row ────────────────────────────────────────────
    subject_name = sessions.first().subject.name if sessions.first() else "All Subjects"
    section_name = sessions.first().section.name if sessions.first() else "All Sections"
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    title_text = f"ATTENDANCE REGISTER — {subject_name.upper()} | Section {section_name} | {date_range}"

    total_cols = 6 + sessions.count() + 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=13)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 35

    # ── Sub info Row ─────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    info_text = f"Faculty: {request.user.get_full_name()} | Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')} | Total Sessions: {sessions.count()}"
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.fill = subheader_fill
    info_cell.font = white_normal
    info_cell.alignment = center
    ws.row_dimensions[2].height = 22

    # ── Empty row ────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Column Headers ────────────────────────────────────────
    header_row = 4
    fixed_headers = ["#", "Enrollment No.", "Student Name", "Branch", "Semester", "Section"]
    for col, h in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border

    # Session date headers
    session_list = list(sessions)
    for i, session in enumerate(session_list):
        col = 7 + i
        date_str = session.date.strftime("%d %b")
        day_str = session.date.strftime("%a")
        cell = ws.cell(row=header_row, column=col, value=f"{date_str}\n{day_str}\n{session.subject.code}")
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Summary headers
    summary_start = 7 + len(session_list)
    for i, h in enumerate(["Total Present", "Total Absent", "Attendance %"]):
        col = summary_start + i
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = subheader_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.row_dimensions[header_row].height = 45

    # ── Fixed column widths ───────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    # ── Student Rows ──────────────────────────────────────────
    for idx, student in enumerate(students):
        row = header_row + 1 + idx
        row_fill = alt_row_fill if idx % 2 == 0 else PatternFill(fill_type=None)

        try:
            profile = student.student_profile
            enrollment = profile.enrollment_number
            branch = profile.branch
            semester = profile.semester
            section = profile.section
        except:
            enrollment = "N/A"
            branch = "N/A"
            semester = "N/A"
            section = "N/A"

        # Fixed columns
        fixed_values = [idx + 1, enrollment, student.get_full_name(), branch, semester, section]
        for col, val in enumerate(fixed_values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.font = dark_bold if col == 3 else Font(size=10)
            cell.alignment = center if col != 3 else left
            cell.border = thin_border

        # Attendance per session
        total_present = 0
        total_absent = 0
        for i, session in enumerate(session_list):
            col = 7 + i
            try:
                record = AttendanceRecord.objects.get(session=session, student=student)
                if record.status == 'present':
                    cell = ws.cell(row=row, column=col, value="P")
                    cell.fill = present_fill
                    cell.font = present_font
                    total_present += 1
                else:
                    cell = ws.cell(row=row, column=col, value="A")
                    cell.fill = absent_fill
                    cell.font = absent_font
                    total_absent += 1
            except AttendanceRecord.DoesNotExist:
                cell = ws.cell(row=row, column=col, value="—")
                cell.fill = row_fill
                cell.font = Font(color="999999", size=10)
            cell.alignment = center
            cell.border = thin_border

        # Summary columns
        total_classes = total_present + total_absent
        percentage = round((total_present / total_classes * 100), 1) if total_classes > 0 else 0

        present_cell = ws.cell(row=row, column=summary_start, value=total_present)
        present_cell.fill = present_fill
        present_cell.font = present_font
        present_cell.alignment = center
        present_cell.border = thin_border

        absent_cell = ws.cell(row=row, column=summary_start + 1, value=total_absent)
        absent_cell.fill = absent_fill
        absent_cell.font = absent_font
        absent_cell.alignment = center
        absent_cell.border = thin_border

        pct_cell = ws.cell(row=row, column=summary_start + 2, value=f"{percentage}%")
        if percentage >= 75:
            pct_cell.fill = present_fill
            pct_cell.font = present_font
        else:
            pct_cell.fill = absent_fill
            pct_cell.font = absent_font
        pct_cell.alignment = center
        pct_cell.border = thin_border

        ws.row_dimensions[row].height = 20

    # ── Legend Row ────────────────────────────────────────────
    legend_row = header_row + len(list(students)) + 2
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=total_cols)
    legend_cell = ws.cell(row=legend_row, column=1, value="LEGEND:   P = Present   |   A = Absent   |   — = No Record   |   Green % = Above 75% (Safe)   |   Red % = Below 75% (At Risk)")
    legend_cell.fill = summary_fill
    legend_cell.font = Font(color="4F46E5", bold=True, size=10)
    legend_cell.alignment = center
    ws.row_dimensions[legend_row].height = 22

    # Freeze top rows
    ws.freeze_panes = "G5"

    # ── Return file ────────────────────────────────────────────
    filename = f"Attendance_{subject_name.replace(' ', '_')}_{section_name}_{datetime.now().strftime('%d%b%Y')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_timetable(request):
    user = request.user
    try:
        profile = user.student_profile
        sections = user.sections.all()
        timetable = Timetable.objects.filter(
            section__in=sections
        ).order_by('day', 'start_time').select_related('subject', 'subject__faculty', 'section')
    except:
        return Response([])

    data = []
    for t in timetable:
        data.append({
            'id': t.id,
            'day': t.day,
            'subject_name': t.subject.name,
            'subject_code': t.subject.code,
            'faculty_name': t.subject.faculty.get_full_name() if t.subject.faculty else 'TBA',
            'section': t.section.name,
            'start_time': t.start_time.strftime('%I:%M %p'),
            'end_time': t.end_time.strftime('%I:%M %p'),
        })
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_faculty_timetable(request):
    user = request.user
    subjects = Subject.objects.filter(faculty=user)
    timetable = Timetable.objects.filter(
        subject__in=subjects
    ).order_by('day', 'start_time').select_related('subject', 'section')

    data = []
    for t in timetable:
        data.append({
            'id': t.id,
            'day': t.day,
            'subject_name': t.subject.name,
            'subject_code': t.subject.code,
            'section': t.section.name,
            'semester': t.section.semester,
            'start_time': t.start_time.strftime('%I:%M %p'),
            'end_time': t.end_time.strftime('%I:%M %p'),
        })
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_students_attendance_summary(request):
    if request.user.role != 'faculty':
        return Response({'error': 'Only faculty can access this'}, status=status.HTTP_403_FORBIDDEN)

    section_id = request.query_params.get('section_id')
    subject_id = request.query_params.get('subject_id')

    if not section_id:
        return Response({'error': 'section_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    section = get_object_or_404(Section, id=section_id)
    students = section.students.filter(role='student').order_by('student_profile__enrollment_number')

    if subject_id:
        subjects = Subject.objects.filter(id=subject_id, faculty=request.user)
    else:
        subjects = Subject.objects.filter(faculty=request.user)

    data = []
    for student in students:
        try:
            profile = student.student_profile
            enrollment = profile.enrollment_number
            branch = profile.branch
            semester = profile.semester
        except:
            enrollment = "N/A"
            branch = "N/A"
            semester = "N/A"

        subject_wise = []
        overall_present = 0
        overall_total = 0

        for subject in subjects:
            total_classes = AttendanceSession.objects.filter(
                subject=subject, section=section
            ).count()
            total_present = AttendanceRecord.objects.filter(
                student=student,
                session__subject=subject,
                session__section=section,
                status='present'
            ).count()
            total_absent = total_classes - total_present
            percentage = round((total_present / total_classes * 100), 1) if total_classes > 0 else 0
            overall_present += total_present
            overall_total += total_classes
            subject_wise.append({
                'subject_id': subject.id,
                'subject_name': subject.name,
                'subject_code': subject.code,
                'total_classes': total_classes,
                'total_present': total_present,
                'total_absent': total_absent,
                'percentage': percentage,
            })

        overall_percentage = round((overall_present / overall_total * 100), 1) if overall_total > 0 else 0
        data.append({
            'student_id': student.id,
            'name': student.get_full_name(),
            'username': student.username,
            'enrollment_number': enrollment,
            'branch': branch,
            'semester': semester,
            'overall_present': overall_present,
            'overall_total': overall_total,
            'overall_percentage': overall_percentage,
            'subject_wise': subject_wise,
            'is_at_risk': overall_percentage < 75 and overall_total > 0,
        })

    return Response({
        'section': f"Section {section.name} - Sem {section.semester}",
        'total_students': len(data),
        'students': data
    })